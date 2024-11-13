import sys
from io import BytesIO
from lifelines import CoxPHFitter
import joblib
import numpy as np
from kivy.app import App
from kivy.uix.image import Image
from kivy.graphics.texture import Texture

from kivy.uix.popup import Popup
from kivy.uix.widget import Widget
from matplotlib import pyplot as plt
from sklearn.ensemble import GradientBoostingClassifier
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.filechooser import FileChooserIconView
from kivy.core.window import Window
from kivy.uix.scrollview import ScrollView
import pandas as pd
import os
import ctypes
from openpyxl import load_workbook
from kivy.core.window import Window
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.label import Label
from kivy.graphics import Rotate
from kivy.uix.label import Label
from kivy.graphics import PushMatrix, PopMatrix, Rotate
from PIL import Image as PILImage
import io

#Window.clearcolor = (0.9, 0.95, 1, 1)
Window.clearcolor = (0.62, 0.72, 0.73, 1)

# Пример задания параметров для кнопок
button_style = {
    "background_color": (0.15, 0.21, 0.44, 1),  # Постоянный голубой цвет кнопки
    "background_normal": "",  # Убирает эффект изменения цвета при нажатии
    "color": (1, 1, 1, 1),  # Белый текст
    "font_size": 22,
    "size_hint": (None, None),
    "size": (200, 50)
}


# Путь к директории с моделями
model_dir = os.path.join(os.getcwd())

# Добавляем модельную директорию в sys.path с использованием dir()
if model_dir not in sys.path:
    sys.path.append(model_dir)

# Проверим содержимое директории с моделями через dir()
print("Содержимое директории с моделями:", os.listdir(model_dir))

# Попробуем загрузить модели
try:
    model_2y = joblib.load(os.path.join(model_dir, "gb_classifier_model.pkl"))  # 2-year survival model
    model_5y = joblib.load(os.path.join(model_dir, "gb_classifier_model_5.pkl"))  # 5-year survival model
    cph = joblib.load(os.path.join(model_dir, "cox_model.pkl"))  # cox survival model
    print("Модели успешно загружены.")
except FileNotFoundError:
    print("Модели не найдены. Убедитесь, что файлы 'gb_classifier_model.pkl' и 'gb_classifier_model_5.pkl' находятся в директории.")
except Exception as e:
    print(f"Произошла ошибка при загрузке моделей: {e}")


class MainScreen(Screen):
    def __init__(self, **kwargs):
        super(MainScreen, self).__init__(**kwargs)
        Window.set_icon(os.path.join(model_dir, "1.png"))

        # Создаем AnchorLayout для центрирования всех элементов
        layout = BoxLayout(orientation='vertical', padding=[525, 500, 0, 350], spacing=20)

        # Добавляем изображение
        image = Image(source=os.path.join(model_dir, "3.png"), size_hint=(None, None), size=(350, 350))
        layout.add_widget(image)


        # Контейнер для кнопок
        button_container = BoxLayout(orientation='vertical',  padding=[-265, 50, 0, -80], spacing=20, size_hint=(None, None), size=(200, 100))

        # Создаем кнопки
        btn_load_file = Button(text="Load file", **button_style)
        btn_manual_input = Button(text="Manual input", **button_style)

        # Привязываем обработчики к кнопкам
        btn_load_file.bind(on_press=self.load_file)
        btn_manual_input.bind(on_press=self.manual_input)

        # Добавляем кнопки в контейнер
        button_container.add_widget(btn_load_file)
        button_container.add_widget(btn_manual_input)

        button_container.size_hint_x = None
        button_container.pos_hint = {'center_x': 0.5}
        layout.add_widget(button_container)

        # Добавляем layout на экран
        self.add_widget(layout)

    def load_file(self, instance):
        self.manager.current = 'filechooser'

    def manual_input(self, instance):
        self.manager.current = 'manualinput'


class FileChooserScreen(Screen):
    def __init__(self, **kwargs):
        super(FileChooserScreen, self).__init__(**kwargs)

        # Основной вертикальный layout
        layout = BoxLayout(orientation='vertical', padding=[50, 50, 50, 50], spacing=30)

        # Горизонтальный контейнер для заголовка с выравниванием слева и отступом
        title_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=40)
        title = Label(text="Select file to upload:", font_size=30, bold=True,
                      color=(0, 0, 0, 1), halign='left', valign='middle', italic=True)
        title.text_size = (1100, None)  # Установка ширины для выравнивания по левому краю
        title_layout.add_widget(title)

        # Пустой отступ перед FileChooserIconView
        title_layout.add_widget(Widget(size_hint_x=0.2))
        layout.add_widget(title_layout)

        # Виджет выбора файла с начальным путем
        self.filechooser = FileChooserIconView(path=os.getcwd(), size_hint=(1, None), height=750)
        layout.add_widget(self.filechooser)

        # Горизонтальный layout для кнопок "Назад" и "Открыть"
        button_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=50, spacing=20)
        btn_back = Button(text="Back", **button_style)
        btn_back.bind(on_press=self.go_back)
        button_layout.add_widget(btn_back)

        btn_open = Button(text="Open", **button_style)
        btn_open.bind(on_press=self.open_file)
        button_layout.add_widget(btn_open)

        layout.add_widget(button_layout)
        self.add_widget(layout)

    def go_back(self, instance):
        # Переход на главный экран
        self.manager.current = 'main'

    def open_file(self, instance):
        selected = self.filechooser.selection
        if selected:
            file_path = selected[0]
            if file_path.endswith('.xlsx'):
                # Загрузка данных из файла и переход к экрану просмотра
                self.manager.get_screen('dataview').display_data(file_path)
                self.manager.current = 'dataview'


class ManualInputScreen(Screen):
    def __init__(self, **kwargs):
        super(ManualInputScreen, self).__init__(**kwargs)

        layout = GridLayout(cols=2, spacing=10, padding=40)

        # Поля для ввода данных с подсказками
        self.inputs = {}

        fields = [
            ("Full name:", "John Doe"),
            ("Sex:", "1 - man, 2 - female"),
            ("Variant:", "1-4"),
            ("Synchronous/Metachronous tumors:", "1 - synchronous, 2 - metachronous"),
            ("Status ECOG:", "0-3"),
            ("MSKCC:", "1-3"),
            ("Differentiation:", "1-3"),
            ("Number of metastases:", "1 - solitary, 2 - single, 3 - multiple"),
            ("Radiotherapy:", "1 - yes, 0 - no"),
            ("Nexavar in first-line therapy:", "1 - yes, 0 - no"),
            ("Intermittency of treatment:", "1 - yes, 0 - no"),
            ("Metastatic surgery:", "1 - yes, 0 - no"),
            ("Nephrectomy:", "1 - yes, 0 - no"),
            ("Progress code:", "1-4"),
        ]

        for label_text, hint_text in fields:
            # Метка для поля
            label = Label(
                text=label_text,
                font_size=22,
                color=(0, 0, 0, 1),  # Черный текст
                halign='left',
                valign='middle', # Вертикальное выравнивание посередине
                bold=True

            )
            label.bind(size=label.setter('text_size'))  # Позволяет тексу переносииться по ширине
            layout.add_widget(label)

            input_field = TextInput(
                hint_text=hint_text,
                hint_text_color=(0.5, 0.5, 0.5, 0.5),
                foreground_color=(0, 0, 0, 1),  # Черный цвет текста
                background_color=(1, 1, 1, 1),  # Белый фон для ввода
                font_size=22,
                multiline=False,
                cursor_color=(0, 0, 0, 1),  # Черный цвет курсора
                border=(1, 1, 1, 1),


            )
            self.inputs[label_text] = input_field
            layout.add_widget(input_field)

        back_button = Button(text="Back",  **button_style)
        back_button.bind(on_press=self.go_back)
        layout.add_widget(back_button)

        btn_save = Button(text="Save", **button_style)
        btn_save.bind(on_press=self.save_data)
        layout.add_widget(btn_save)

        self.add_widget(layout)

    def save_data(self, instance):
        # Сохранение данных из полей
        data = {field: self.inputs[field].text for field in self.inputs}

        cleaned_data = {field.rstrip(':'): value for field, value in data.items()}
        df = pd.DataFrame([cleaned_data])

        '''file_path = "base.xlsx"

        try:
            # Загружаем существующий файл
            book = load_workbook(file_path)
            writer = pd.ExcelWriter(file_path, engine="openpyxl", mode="a")
            writer.book = book

            # Определяем существующий лист и добавляем данные в конец
            startrow = writer.sheets["Sheet1"].max_row
            df.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name="Sheet1")
            writer.save()
            writer.close()

        except FileNotFoundError:
            # Если файла нет, создаем новый
            df.to_excel(file_path, index=False, sheet_name="Sheet1")'''

            # Очистка полей ввода после сохранения данных
        for input_field in self.inputs.values():
            input_field.text = ""

            # Отображение данных на экране DataView
        self.manager.get_screen('dataview').display_data(df)
        self.manager.current = 'dataview'

    def go_back(self, instance):
        # Очистка всех полей ввода
        for input_field in self.inputs.values():
            input_field.text = ""

        # Возвращение на главный экран
        self.manager.current = 'main'

class RotatedLabel(Label):
    def __init__(self, **kwargs):
        self.angle = kwargs.pop('angle', 0)  # Устанавливаем угол поворота по умолчанию
        super(RotatedLabel, self).__init__(**kwargs)

        # Включаем выравнивание по левому краю и отступ по Y
        self.halign = 'center'
        self.valign = 'middle'
        self.text_size = (self.width, None)  # Ограничиваем текст по ширине, чтобы работало выравнивание
        self.padding_y = 30  # Добавляем отступ сверху

        # Добавляем команды для поворота и восстановления матрицы
        with self.canvas.before:
            self.push_matrix = PushMatrix()
            self.rotation = Rotate(angle=self.angle, origin=self.center)
        with self.canvas.after:
            self.pop_matrix = PopMatrix()

        # Обновляем центр поворота при изменении размера и позиции
        self.bind(pos=self.update_rotation, size=self.update_rotation)
        self.bind(width=self.update_text_size)

    def update_rotation(self, *args):
        self.rotation.origin = self.center  # Устанавливаем центр поворота

    def update_text_size(self, *args):
        self.text_size = (self.width, None)


class DataViewScreen(Screen):
    def __init__(self, **kwargs):
        super(DataViewScreen, self).__init__(**kwargs)

        # Основной вертикальный layout
        layout = BoxLayout(orientation='vertical', padding=40, spacing=100)

        # Прокручиваемая область для таблицы данных (по горизонтали и вертикали)
        self.scroll_view = ScrollView(size_hint=(1, None), size=(1200, 750), do_scroll_x=True, do_scroll_y=True, bar_width=10)
        self.data_layout = GridLayout(cols=1, spacing=50,size_hint_y=None, size_hint_x=None)
        self.scroll_view.add_widget(self.data_layout)

        # Кнопки "Получить результат" и "Отмены"
        buttons_layout = BoxLayout(orientation='horizontal', size_hint_y=None, height=50, spacing=50)

        btn_analyze = Button(text="Get Prediction", **button_style)
        btn_analyze.bind(on_press=self.get_prediction_for_all)
        buttons_layout.add_widget(btn_analyze)

        btn_cancel = Button(text="Back",  **button_style)
        btn_cancel.bind(on_press=self.go_back_to_main)
        buttons_layout.add_widget(btn_cancel)

        # Добавление виджетов на главный layout
        layout.add_widget(self.scroll_view)
        layout.add_widget(buttons_layout)
        self.add_widget(layout)

    def go_back_to_main(self, instance):
        # Очистка таблицы данных
        self.data_layout.clear_widgets()
        self.manager.current = 'main'

    def display_data(self, data):
        # Очистка предыдущих данных
        self.data_layout.clear_widgets()

        if isinstance(data, str):
            if os.path.exists(data):  # Проверяем, существует ли файл
                try:
                    df = pd.read_excel(data, header=0)
                    print("Данные загружены из файла:", df.head())  # Отладочное сообщение
                except Exception as e:
                    print(f"Ошибка при загрузке данных из файла: {e}")
                    error_label = Label(text="Error loading file", font_size=16)
                    self.data_layout.add_widget(error_label)
                    return
            else:
                error_label = Label(text="File not found", font_size=16)
                self.data_layout.add_widget(error_label)
                return
        elif isinstance(data, pd.DataFrame):
            df = data
            print("DataFrame загружен напрямую:", df.head())  # Отладочное сообщение
        else:
            error_label = Label(text="Invalid data type", font_size=16)
            self.data_layout.add_widget(error_label)
            return

        self.data = df

        # Устанавливаем количество столбцов
        self.data_layout.cols = len(df.columns)
        self.data_layout.height = 70 * (len(df) + 1)

        # Динамический расчет ширины столбцов
        column_widths = []
        max_width = 150  # Максимальная ширина столбца для предотвращения чрезмерного расширения

        for column in df.columns:
            # Получаем максимальную длину строки в столбце (с учетом заголовка)
            max_len = max(df[column].apply(lambda x: len(str(x))).max(), len(str(column)))
            # Ограничиваем ширину столбца
            column_widths.append(min(max_len * 15, max_width))  # Умножаем на коэффициент и ограничиваем max_width

        # Устанавливаем общую ширину gridLayout с учетом расстояния между столбцами
        self.data_layout.width = sum(column_widths) + 150 + (len(df.columns) - 1) * 20  # Добавляем дополнительные отступы

        # Устанавливаем расстояние между столбцами
        self.data_layout.spacing = 30  # Устанавливаем расстояние между столбцами (можно увеличить по желанию)

        # Добавляем заголовки столбцов в layout с отступами
        for column, col_width in zip(df.columns, column_widths):
            header_label = RotatedLabel(
                text=column,  # Добавляем пробелы до и после текста
                bold=True,
                font_size=22,
                size_hint_y=80,
                height=20,
                angle=0,
                color=(0, 0, 0, 1),
                size_hint_x=None,
                width=col_width,
                halign='center',
                valign='top',
            )
            self.data_layout.add_widget(header_label)

        # Добавляем строки данных
        for row in df.itertuples(index=False):
            for i, value in enumerate(row):
                # Устанавливаем ширину ячеек, соответствующую ширине столбца
                value_label = Label(
                    text=str(value),
                    font_size=20,
                    size_hint_y=None,
                    height=40,
                    size_hint_x=None,
                    width=column_widths[i],
                    halign='center',  # Выравнивание по горизонтали по центру
                    valign='middle'  # Выравнивание по вертикали по центру
                )
                self.data_layout.add_widget(value_label)

                # Это обновит все виджеты на экране

    def get_prediction_for_all(self, instance):
        columns_name = [
            'Intermittency', 'Number of metastases', 'Status', '1Synchron/2Metachron', 'Nephrectomy',
            'MSKCC', 'Differentiation', 'Ray coding', 'Nexavar', 'Metastatic surgery', 'Variant',
            'Sex 1 - m 2 - w', 'Progress code'
        ]
        field_mapping = {
            "Sex": "Sex 1 - m 2 - w",
            "Variant": "Variant",
            "Synchronous/Metachronous tumors": "1Synchron/2Metachron",
            "Status ECOG": "Status",
            "MSKCC": "MSKCC",
            "Differentiation": "Differentiation",
            "Number of metastases": "Number of metastases",
            "Radiotherapy": "Ray coding",
            "Nexavar in first-line therapy": "Nexavar",
            "Intermittency of treatment": "Intermittency",
            "Metastatic surgery": "Metastatic surgery",
            "Nephrectomy": "Nephrectomy",
            "Progress code": "Progress code"
        }

            # Создаем содержимое для Popup с предсказаниями
        prediction_content = BoxLayout(orientation='vertical', padding=(10, 10), spacing=10, size_hint_y=None)
        prediction_content.bind(minimum_height=prediction_content.setter('height'))

        scroll_view = ScrollView(size_hint=(1, 1), bar_width=10)  # Устанавливаем размер ScrollView
        scroll_view.add_widget(prediction_content)  # Добавляем содержимое в ScrollView

        # Проходим по каждой строке данных
        for idx, row in self.data.iloc[0:].iterrows():
            try:
                # Логируем входные данные для проверки
                print(f"Данные для пациента {idx + 1 }: {row.to_dict()}")  # Логируем строку

                # Подготовка данных для подачи в модель с учетом маппинга
                model_input = []

                # Перебираем поля в соответствии с полями из columns_name
                for col in columns_name:
                    # Ищем соответствующее имя столбца в field_mapping
                    original_field = next((key for key, value in field_mapping.items() if value == col), None)

                    if original_field:
                        value = row.get(original_field)
                        model_input.append(value)


                # Логируем подготовленные данные для модели
                print(f"Подготовленные данные для модели: {model_input}")

                # Подготовка данных для подачи в модель
                test_data = np.array([model_input])

                # Проверим форму тестовых данных перед подачей в модель
                print(f"Форма данных для модели: {test_data.shape}")

                # Получаем предсказания для 2 и 5 лет
                prob_2y = model_2y.predict_proba(test_data)[0, 1]
                prob_5y = model_5y.predict_proba(test_data)[0, 1]
                risk = cph.predict_partial_hazard(test_data)

                patient_name = row.get("Full name")

                # Форматируем результат
                name_label = Label(text=f"Patient {patient_name}:", font_size=25, size_hint_y=None, height=30,
                                   halign='left')
                name_label.text_size = (1127, None)
                prediction_content.add_widget(name_label)

                survival_2y_label = Label(text=f"2-year survival = {prob_2y:.2f}", font_size=22, size_hint_y=None,
                                          height=30, halign='left')
                survival_2y_label.text_size = (950, None)
                prediction_content.add_widget(survival_2y_label)

                survival_5y_label = Label(text=f"5-year survival = {prob_5y:.2f}", font_size=22, size_hint_y=None,
                                          height=30, halign='left')
                survival_5y_label.text_size = (950, None)
                prediction_content.add_widget(survival_5y_label)

                try:
                    survival_function = cph.predict_survival_function(test_data)

                    # Если survival_function имеет правильную форму, построим график
                    if not survival_function.empty:
                        plt.figure(figsize=(7, 5))
                        plt.plot(survival_function.index, survival_function.values, linestyle='-',
                                 color='b', linewidth=2)
                        plt.title(f"Survival Curve for {patient_name}")
                        plt.xlabel("Time")
                        plt.ylabel("Survival Probability")
                        plt.grid(True)

                        # Сохраним график в буфер и конвертируем в изображение для отображения в Kivy
                        buf = BytesIO()
                        plt.savefig(buf, format='jpeg')
                        buf.seek(0)
                        plt.close()

                        pil_img = PILImage.open(buf)
                        pil_img = pil_img.convert("RGBA")
                        img_data = np.array(pil_img)

                        img_data = np.flipud(img_data)

                        texture = Texture.create(size=(img_data.shape[1], img_data.shape[0]), colorfmt='rgba')
                        texture.blit_buffer(img_data.tobytes(), colorfmt='rgba', bufferfmt='ubyte')

                        img = Image(texture=texture, size_hint=(None, None), size=(700, 500))
                        prediction_content.add_widget(img)

                        lab = Label(text=f" ", font_size=22,
                                                  size_hint_y=None,
                                                  height=50, halign='left')
                        lab.text_size = (950, None)
                        prediction_content.add_widget(lab)

                    else:
                        print(f"Ошибка: данные для кривой выживаемости отсутствуют для пациента {patient_name}")

                except Exception as e:
                    print(f"Ошибка при обработке пациента {idx + 1}: {e}")
                    error_label = Label(text=f"Error for patient {idx + 1}", font_size=16, size_hint_y=None, height=30, halign='left', valign='middle')
                    error_label.text_size = (2000, None)  # Выровнять по ширине
                    prediction_content.add_widget(error_label)

            except Exception as e:
                print(f"Ошибка при обработке пациента {idx + 1}: {e}")
                error_label = Label(text=f"Error for patient {idx + 1}", font_size=16, size_hint_y=None, height=30, halign='left',valign='middle')
                error_label.text_size = (2000, None)  # Выровнять по ширине
                prediction_content.add_widget(error_label)

        main_layout = BoxLayout(orientation='vertical', spacing=10)
        main_layout.add_widget(scroll_view)

        # Фиксируем кнопку Close внизу окна
        close_button = Button(text="Close", **button_style)
        close_button.bind(on_press=lambda x: popup.dismiss())
        main_layout.add_widget(close_button)

        # Создаем Popup
        popup = Popup(title="Prediction Results",title_size=28, content=main_layout, size_hint=(None, None), size=(1200, 1000))
        popup.open()


    def go_back_to_main(self, instance):
        self.data_layout.clear_widgets()
        self.manager.current = 'main'


class MyApp(App):
    def build(self):
        # Установка начального размера окна
        Window.size = (800, 600)
        self.title = "Kidney Survive"

        # Создаем менеджер экранов и добавляем экраны
        sm = ScreenManager()
        sm.add_widget(MainScreen(name='main'))
        sm.add_widget(FileChooserScreen(name='filechooser'))
        sm.add_widget(ManualInputScreen(name='manualinput'))
        sm.add_widget(DataViewScreen(name='dataview'))

        # Центрируем окно
        self.center_window()

        return sm

    def center_window(self):
        # Получаем размер экрана
        screen_width, screen_height = Window.system_size
        # Устанавливаем окно в центр
        Window.left = (screen_width - 100) // 2
        Window.top = (screen_height - 200 ) // 2


if __name__ == '__main__':
    MyApp().run()
